from matplotlib.pylab import plt
import numpy as np
import pandas as pd
from sklearn.metrics import confusion_matrix, classification_report
import seaborn as sns
from sklearn.decomposition import PCA
from sklearn.manifold import TSNE
import tensorflow as tf
import warnings
from sklearn.exceptions import UndefinedMetricWarning
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import tempfile  # To create a temporary file
import os


def save_trained_model(vae, optimizer, model_path,model_name = "", latent_dim = "", beta = "", n_rows_train = "", time = "", AWS = False, s3 = None, BUCKET = ""):
    if AWS:
        s3_key = f'Models/BEST_{model_name}_LD{latent_dim}_Beta{beta}_NT{n_rows_train}_{time}.keras'
        vae.compile(optimizer = optimizer)
        vae.save("vae_model.keras")
        # Upload to S3
        s3.upload_file("vae_model.keras", BUCKET, s3_key)
        print(f"Model saved to S3: s3://{BUCKET}/{s3_key}")
    else:
        vae.compile(optimizer = optimizer)
        vae.save(model_path)
    

# PLOT LOSS
def plot_loss_curve(epoch_losses, val_losses, epochs, latent_dim, time,n_rows_train, show_val = False, AWS = False, s3 = None, BUCKET = ""):
    # Plotting training and validation loss curves
    plt.figure(figsize=(8, 6))
    plt.plot(range(epochs), epoch_losses, label='Training Loss', color='blue')
    if show_val:
        plt.plot(range(epochs), val_losses, label='Validation Loss', color='red', linestyle='dashed')
    
    plt.xlabel('Epochs')
    plt.ylabel('Loss')
    plt.title('VAE Training & Validation Loss Curve')
    plt.legend()
    plt.grid(True)

    # Save
    if AWS:
        plot_name = "VAE_MS"
        s3_key = f"Plots/{plot_name}_LD{latent_dim}_EP{epochs}_NT{n_rows_train}_{time}.png"

        # Create a temporary file
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_file:
            plot_path = tmp_file.name
            plt.savefig(plot_path)  # Save plot locally

        # Upload to S3
        s3.upload_file(plot_path, BUCKET, s3_key)
        print(f"✅ Plot saved to S3: s3://{BUCKET}/{s3_key}")

        # Remove local temp file
        os.remove(plot_path)
    else:
        plot_name = "VAE_MS"
        plot_path = f"./Resources/Plots/{plot_name}_LD{latent_dim}_EP{epochs}_NT{n_rows_train}_{time}.png"
        plt.savefig(plot_path)

    # Show
    plt.show()


##
def get_confusion_matrix(true_labels, predicted, latent_dim , epochs,time, n_rows_train,categories = ["normal" , "attack"], AWS = False, s3 = None, BUCKET = ""):
    cm = confusion_matrix(true_labels, predicted)
    annot = np.where(cm != 0, cm, '')
    sns.heatmap(cm, annot=annot, fmt="", cmap="Blues", cbar=True,
                xticklabels= categories, yticklabels=categories)
    
    plt.xticks(rotation=45, ha="right", fontsize=10)
    plt.title(f"Confusion Matrix", fontsize=15)
    plt.xlabel("Predicted Labels", fontsize=12)
    plt.ylabel("True Labels", fontsize=12)

    if AWS:
        # Create a temporary file
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_file:
            plot_path = tmp_file.name
            plt.savefig(plot_path)  # Save to temp file

        # Define S3 path
        s3_key = f"Plots/Conf_VAE_MS_LD{latent_dim}_EP{epochs}_T{time}.png"

        # Upload to S3
        s3.upload_file(plot_path, BUCKET, s3_key)
        print(f"Plot saved to S3: s3://{BUCKET}/{s3_key}")

        # Remove local temp file
        os.remove(plot_path)
    else:
        plot_name ="Conf_VAE_MS"
        plot_name = f"./Resources/Plots/{plot_name}_LD{latent_dim}_EP{epochs}_NT{n_rows_train}_{time}.png"
        plt.savefig(plot_name)
    
    plt.show()


def plot_pca(normal_dataframe, mixed_dataframe):
    train_features = np.vstack(normal_dataframe['features'].values)

    # Filter only attack samples (type = 1)
    attack_samples = mixed_dataframe[mixed_dataframe['type'] == 1]
    
    if attack_samples.shape[0] > 0:
        attack_features = np.vstack(attack_samples['features'].apply(np.array).values)
    else:
        attack_features = None  # No attack data
        print("0 Attack Logs in Test Set")

    pca = PCA(n_components=2)
    normal_proj = pca.fit_transform(train_features)  

    plt.figure(figsize=(8, 6))
    plt.scatter(normal_proj[:, 0], normal_proj[:, 1], c='blue', alpha=0.5, label="Normal", marker = 'o')

    if attack_features is not None:
        attack_proj = pca.transform(attack_features)
        plt.scatter(attack_proj[:, 0], attack_proj[:, 1], c='red', alpha=0.25, label="Attack", marker = 'x')

    plt.xlabel("PCA Component 1")
    plt.ylabel("PCA Component 2")
    plt.legend()
    plt.title("PCA Projection of CAN Data")
    plt.show()


def plot_tsne(normal_dataframe, mixed_dataframe):
    train_features = np.vstack(normal_dataframe['features'].values)

    # Filter only attack samples (type = 1)
    attack_samples = mixed_dataframe[mixed_dataframe['type'] == 1]
    
    if attack_samples.shape[0] > 0:
        attack_features = np.vstack(attack_samples['features'].apply(np.array).values)
        combined_features = np.vstack([train_features, attack_features])  # Combine normal + attack data
        labels = np.hstack([np.zeros(train_features.shape[0]), np.ones(attack_features.shape[0])])  # 0 for normal, 1 for attack
    else:
        combined_features = train_features
        labels = np.zeros(train_features.shape[0])
        print("0 Attack Logs in Test Set")

    # Apply t-SNE
    tsne = TSNE(n_components=2, random_state=42, perplexity=30, n_iter=1000)
    tsne_proj = tsne.fit_transform(combined_features)

    # Plot
    plt.figure(figsize=(8, 6))
    plt.scatter(tsne_proj[labels == 0, 0], tsne_proj[labels == 0, 1], c='blue', alpha=0.5, label="Normal")
    
    if 1 in labels:
        plt.scatter(tsne_proj[labels == 1, 0], tsne_proj[labels == 1, 1], c='red', alpha=0.5, label="Attack")

    plt.xlabel("t-SNE Dimension 1")
    plt.ylabel("t-SNE Dimension 2")
    plt.legend()
    plt.title("t-SNE Projection of CAN Data")
    plt.show()


def get_latent_representations_label(vae, dataset, latent_dim, beta ,n_critic,gamma,time,epoch,name = "",type='PCA', save = False, AWS = False, s3 = None, BUCKET = ""):    
    latent_representations = []
    labels = []  # Collect labels if available

    for batch in dataset:
        if isinstance(batch, (tuple, list)):  
            batch, batch_labels = batch  
            labels.append(batch_labels.numpy())  
        else:
            batch_labels = None  # 

        _, mu, _ = vae(batch, n_samples=1, latent_only = True)
        latent_representations.append(mu.numpy())

    latent_representations = np.concatenate(latent_representations, axis=0)

    if labels:
        labels = np.concatenate(labels, axis=0)  
    else:
        labels = None  

    if type == 'PCA':
        reducer = PCA(n_components=2)
    elif type == 'TSNE':
        reducer = TSNE(n_components=2, random_state=42)
    else:
        raise ValueError("Invalid type. Choose 'PCA' or 'TSNE'.")

    latent_2d = reducer.fit_transform(latent_representations)

    # Plot # TOdo: plot dimension 2,3
    plt.figure(figsize=(10, 8))

    if labels is not None:
        scatter = plt.scatter(latent_2d[:, 0], latent_2d[:, 1], c=labels, cmap='tab10', alpha=0.3)
        plt.legend(*scatter.legend_elements(), title="Labels")
    else:
        plt.scatter(latent_2d[:, 0], latent_2d[:, 1], color='blue', alpha=0.5, label="Unlabeled (Train)")

    plt.title(f"Latent Space Visualization ({type})")
    plt.xlabel("Dimension 1")
    plt.ylabel("Dimension 2")
    plt.legend()

    # Save 
    if save:
        if AWS:
            plot_name = f"{name}_Latent_Space_EPOCH{epoch}_LD{latent_dim}_Beta{beta}_n_critic{n_critic}_gamma{gamma}_time{time}.png"

            # Save plot to a temporary file
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_file:
                temp_path = temp_file.name
                plt.savefig(temp_path)
                temp_file.close()

                # Upload to S3
                s3_key = f"Plots/{plot_name}"  # Change 'plots/' to your desired S3 folder
                s3.upload_file(temp_path, BUCKET, s3_key)

                # Delete the temporary file after uploading
                os.remove(temp_path)

                print(f"Plot saved to S3: s3://{BUCKET}/{s3_key}")
        else:
            plot_name ="Latent_Space"
            plot_name = f"./Resources/Plots/{name}_{plot_name}_EPOCH{epoch}_LD{latent_dim}_Beta{beta}_n_critic{n_critic}_gamma{gamma}_time{time}.png"
            plt.savefig(plot_name)

    plt.show()


def analyze_latent_variance(vae, train, test,device='cuda'):
    latents_train = []
    latents_test = []

    # Collect latent vectors from the training dataset
    for x in train:
        _, mu, logvar = vae(x, n_samples=1, latent_only=True)
        z = vae.reparameterize(mu, logvar, n_samples=1)
        z = tf.squeeze(z, axis=0)  # Remove sample dimension
        latents_train.append(z.numpy())

    latents_train = np.concatenate(latents_train, axis=0)
    train_variances = np.var(latents_train, axis=0)

    # Collect latent vectors from the test dataset
    for x in test:
        _, mu, logvar = vae(x, n_samples=1, latent_only=True)
        z = vae.reparameterize(mu, logvar, n_samples=1)
        z = tf.squeeze(z, axis=0)  # Remove sample dimension
        latents_test.append(z.numpy())

    latents_test = np.concatenate(latents_test, axis=0)
    test_variances = np.var(latents_test, axis=0)

    # Plotting both train and test latent variances
    plt.figure(figsize=(10, 5))
    latent_dim = len(train_variances)
    plt.bar(np.arange(latent_dim) - 0.2, train_variances, width=0.4, label="Train Variance", alpha=0.7)
    plt.bar(np.arange(latent_dim) + 0.2, test_variances, width=0.4, label="Test Variance", alpha=0.7)
    
    plt.xlabel("Latent Dimension")
    plt.ylabel("Variance")
    plt.title("Latent Space Variance (Train vs Test)")
    plt.legend()
    plt.show()


def analyze_kl_divergence(vae, train, test, device='cuda'):
    kl_train_values = []
    kl_test_values = []

    # Calculate KL divergence for the training set
    for x in train:
        _, mu, logvar = vae(x, n_samples=1, latent_only=True)
        kl = -0.5 * (1 + logvar - tf.square(mu) - tf.exp(logvar))  
        kl_train_values.append(kl.numpy())  # Use .numpy() for TF tensors

    kl_train_values = np.concatenate(kl_train_values, axis=0)
    mean_kl_train = np.mean(kl_train_values, axis=0)

    # Calculate KL divergence for the test set
    for x in test:
        _, mu, logvar = vae(x, n_samples=1, latent_only=True)
        kl = -0.5 * (1 + logvar - tf.square(mu) - tf.exp(logvar))  
        kl_test_values.append(kl.numpy())  # Use .numpy() for TF tensors

    kl_test_values = np.concatenate(kl_test_values, axis=0)
    mean_kl_test = np.mean(kl_test_values, axis=0)

    # Plotting both train and test KL divergence values
    plt.figure(figsize=(10, 5))
    latent_dim = len(mean_kl_train)
    plt.bar(np.arange(latent_dim) - 0.2, mean_kl_train, width=0.4, label="Train KL Divergence", alpha=0.7)
    plt.bar(np.arange(latent_dim) + 0.2, mean_kl_test, width=0.4, label="Test KL Divergence", alpha=0.7)
    
    plt.xlabel("Latent Dimension")
    plt.ylabel("KL Divergence")
    plt.title("KL Divergence per Latent Dimension (Train vs Test)")
    plt.legend()
    plt.show()


def linear_annealing(start, end, step, total_steps):
    """Linearly anneal from start to end over total_steps."""
    return start + (end - start) * tf.minimum(step / total_steps, 1.0)


def save_results_to_excel(model_name, true_labels, predicted_labels, excel_file_path):
    """
    Save confusion matrix and classification report to a specified sheet in an Excel file.

    Parameters:
    - model_name (str): Name of the model (used as the sheet name).
    - true_labels (pd.Series or list): True labels for the data.
    - predicted_labels (pd.Series or list): Predicted labels for the data.
    - excel_file_path (str): Path to the Excel file to save the results.
    """
    # Compute confusion matrix and classification report
    conf_matrix = confusion_matrix(true_labels, predicted_labels)
    
    # Suppress undefined recall warning
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=UndefinedMetricWarning) # Ignores warnings when Test Data does not have any Anomalies
        class_report = classification_report(true_labels, predicted_labels, output_dict=True, zero_division=0)
    
    # Convert classification report to a DataFrame
    class_report_df = pd.DataFrame(class_report).transpose()

    # Load or create an Excel workbook
    try:
        workbook = load_workbook(excel_file_path)  # Load existing file
    except FileNotFoundError:
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove default sheet

    # Remove existing sheet if already present
    if model_name in workbook.sheetnames:
        workbook.remove(workbook[model_name])  
    worksheet = workbook.create_sheet(title=model_name)

    # Write confusion matrix
    worksheet.append(["Confusion Matrix"])
    for row in conf_matrix.tolist():  # Explicitly convert NumPy array to list
        worksheet.append(row)
    
    worksheet.append([])  # Add a blank row

    # Check if there are no '1' labels in true_labels
    if not any(label == 1 for label in true_labels):
        print("⚠️⚠️ No Anomalies present in TESTING SET!! ⚠️⚠️")

    # Write classification report
    worksheet.append(["Performance Report"])
    for row in dataframe_to_rows(class_report_df, index=True, header=True):
        worksheet.append(row)

    # Print results to console
    print("Confusion Matrix:")
    print(conf_matrix)

    print("\nPerformance Report:")
    print(classification_report(true_labels, predicted_labels, zero_division= 0))  # Fixed variable name

    ## Save the workbook
    #workbook.save(excel_file_path)
    print(f"✅ Results saved to sheet '{model_name}' in {excel_file_path}")

def get_s3_client(REGION = 'eu-west-1',BUCKET = 'ml-can-ids-logs', print_files = True):
    import boto3
    
    s3 = boto3.client('s3', region_name= REGION)
    client = boto3.client("sagemaker")
    response = client.describe_notebook_instance(NotebookInstanceName="MLCANIDS")
    print(response["InstanceType"])
    # List objects in the bucket
    response = s3.list_objects_v2(Bucket= BUCKET)

    # Print the file names in the bucket
    if print_files:
        for obj in response.get('Contents', []):
            print(obj['Key'])
    
    return s3


def check_dataset(dataset, dataset_name="Dataset"):
    all_values = []
    
    for batch in dataset:
        # Convert TensorFlow tensors to NumPy
        batch_np = batch.numpy() if isinstance(batch, tf.Tensor) else np.array(batch)
        
        # Flatten batch and store values
        all_values.extend(batch_np.flatten())
    
    # Perform checks
    has_nan = np.isnan(all_values).any()
    has_out_of_bounds = (np.array(all_values) < 0).any() or (np.array(all_values) > 1).any()
    all_values = [lst if isinstance(lst, (list, np.ndarray)) else [lst] for lst in all_values]
    all_same_size = all(len(lst) == len(all_values[0]) for lst in all_values)
    # Final condition combining both checks
    if not all_same_size:
        print("Error: Lists have different sizes.")
    print(f"{dataset_name} - Contains NaN: {has_nan}")
    print(f"{dataset_name} - Contains values <0 or >1: {has_out_of_bounds}")
